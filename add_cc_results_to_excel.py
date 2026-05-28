#!/usr/bin/env python3
"""
Add CubePB/C&C results to the benchmarks Excel.

Expected layout when you run it:

    ./benchmarks.xlsx
    ./OPT/   # extracted results folder, optional
    ./DEC/   # extracted results folder, optional
    ./add_cc_results_to_excel.py

Usage:
    python3 add_cc_results_to_excel.py benchmarks.xlsx

Optional:
    python3 add_cc_results_to_excel.py benchmarks.xlsx --opt-dir OPT --dec-dir DEC --name CubePB --output benchmarks_with_CubePB.xlsx

What it does:
  - Reads result files from OPT/ and/or DEC/.
  - Ignores *.out-* and *.err-* files.
  - Extracts the benchmark ID from the filename prefix, e.g. 106-normalized-xxx.opb -> 106.
  - Extracts Real time from lines like: Real: 148.23, User: ..., Sys: ...
  - Extracts OPT objective from lines like: Global result: OPTIMUM = -34
  - Adds/updates a new solver block:
        OPT sheets: Status / Time / Opt
        DEC sheets: Status / Time
  - Preserves formatting by copying the style from the previous solver block.

Requires:
    pip install openpyxl
"""

from __future__ import annotations

import argparse
import copy
import re
from pathlib import Path
from typing import Dict, Optional, Tuple, Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

Result = Dict[str, Any]


def extract_id_from_filename(path: Path) -> Optional[int]:
    """Extract benchmark ID from a filename. Prefer a leading numeric prefix."""
    name = path.name

    # Strip result suffixes if present, although normally they are ignored.
    name = re.sub(r"\.(out|err)-\d+$", "", name)

    # Strip known benchmark/compression extensions.
    stem = name
    while True:
        new = re.sub(r"\.(xz|gz|bz2|zip|cnf|opb|wcnf|mps|lp|res)$", "", stem, flags=re.I)
        if new == stem:
            break
        stem = new

    # Preferred case: 106-normalized-....opb
    m = re.match(r"^0*(\d+)(?:\D|$)", stem)
    if m:
        return int(m.group(1))

    # Fallback: last numeric token.
    nums = re.findall(r"\d+", stem)
    if nums:
        return int(nums[-1])

    return None


def parse_result_file(path: Path) -> Result:
    text = path.read_text(errors="replace")

    real_matches = re.findall(r"Real:\s*([0-9]+(?:\.[0-9]+)?)", text)
    real = float(real_matches[-1]) if real_matches else None

    maxrss_matches = re.findall(r"MaxRSS_KB:\s*([0-9]+)", text)
    maxrss_kb = int(maxrss_matches[-1]) if maxrss_matches else None

    status = "UNKNOWN"
    opt_value = None

    m = re.search(r"Global result:\s*OPTIMUM\s*=\s*([-+]?\d+(?:\.\d+)?)", text)
    if m:
        status = "OPTIMUM FOUND"
        raw = m.group(1)
        opt_value = float(raw) if "." in raw else int(raw)
    elif re.search(r"\bUNSATISFIABLE\b", text):
        status = "UNSATISFIABLE"
    elif re.search(r"\bSATISFIABLE\b", text):
        status = "SATISFIABLE"

    # Timeout usually produces non-zero status and real time close to the limit.
    if "Command exited with non-zero status" in text and status == "UNKNOWN":
        status = "TIMEOUT/ERROR"

    return {
        "status": status,
        "time": real,
        "opt": opt_value,
        "maxrss_kb": maxrss_kb,
        "file": str(path),
    }


def read_results(results_dir: Path) -> Dict[int, Result]:
    results: Dict[int, Result] = {}
    if not results_dir.exists():
        return results

    for path in results_dir.rglob("*"):
        if not path.is_file():
            continue
        if re.search(r"\.(out|err)-\d+$", path.name):
            continue

        bench_id = extract_id_from_filename(path)
        if bench_id is None:
            continue

        parsed = parse_result_file(path)
        # If duplicates exist, keep the one that has a Real time; otherwise last one wins.
        old = results.get(bench_id)
        if old is None or (old.get("time") is None and parsed.get("time") is not None):
            results[bench_id] = parsed

    return results


def copy_cell_format(src, dst) -> None:
    if src.has_style:
        dst._style = copy.copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy.copy(src.alignment)
    if src.protection:
        dst.protection = copy.copy(src.protection)


def find_id_column(ws: Worksheet) -> int:
    for row in (1, 2):
        for col in range(1, ws.max_column + 1):
            if str(ws.cell(row, col).value).strip().upper() == "ID":
                return col
    # In Selection OPT, A1 is blank but IDs are in column A.
    return 1


def find_existing_block(ws: Worksheet, solver_name: str) -> Optional[int]:
    target = solver_name.strip().lower()
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if isinstance(val, str) and val.strip().lower() == target:
            return col
    return None


def choose_template_start(ws: Worksheet, kind: str) -> int:
    """Choose columns to copy formatting from."""
    if kind == "opt":
        # Prefer last known OPT solver block: Gurobi Status/Time/Opt.
        for col in range(ws.max_column, 0, -1):
            if ws.cell(2, col).value == "Status" and ws.cell(2, col + 1).value == "Time" and ws.cell(2, col + 2).value == "Opt":
                return col
        return 3
    else:
        # Prefer last known DEC solver block: Status/Time.
        for col in range(ws.max_column, 0, -1):
            if ws.cell(2, col).value == "Status" and ws.cell(2, col + 1).value == "Time":
                return col
        return 3


def add_or_update_block(ws: Worksheet, results: Dict[int, Result], kind: str, solver_name: str) -> Tuple[int, int]:
    """
    kind='opt': writes Status/Time/Opt.
    kind='dec': writes Status/Time.
    Returns (matched_rows, start_col).
    """
    ncols = 3 if kind == "opt" else 2
    id_col = find_id_column(ws)
    existing = find_existing_block(ws, solver_name)

    if existing:
        start_col = existing
    else:
        start_col = ws.max_column + 1
        template_start = choose_template_start(ws, kind)

        for offset in range(ncols):
            src_col = template_start + offset
            dst_col = start_col + offset
            ws.column_dimensions[get_column_letter(dst_col)].width = ws.column_dimensions[get_column_letter(src_col)].width

            # Copy row styles down to existing max row.
            for row in range(1, ws.max_row + 1):
                copy_cell_format(ws.cell(row, src_col), ws.cell(row, dst_col))

    # Header row 1.
    ws.cell(1, start_col).value = solver_name
    if ncols > 1:
        try:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + ncols - 1)
        except ValueError:
            # Already merged or overlaps; ignore.
            pass

    # Header row 2.
    headers = ["Status", "Time", "Opt"] if kind == "opt" else ["Status", "Time"]
    for offset, h in enumerate(headers):
        ws.cell(2, start_col + offset).value = h

    matched = 0
    for row in range(3, ws.max_row + 1):
        raw_id = ws.cell(row, id_col).value
        if raw_id is None:
            continue
        try:
            bench_id = int(raw_id)
        except Exception:
            continue

        result = results.get(bench_id)
        if result is None:
            continue

        matched += 1
        ws.cell(row, start_col).value = result.get("status")
        ws.cell(row, start_col + 1).value = result.get("time")
        ws.cell(row, start_col + 1).number_format = "0.00"
        if kind == "opt":
            ws.cell(row, start_col + 2).value = result.get("opt")

    return matched, start_col


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("excel", type=Path, help="Input benchmarks .xlsx file")
    parser.add_argument("--opt-dir", type=Path, default=Path("OPT"), help="OPT results folder")
    parser.add_argument("--dec-dir", type=Path, default=Path("DEC"), help="DEC results folder")
    parser.add_argument("--name", default="CubePB", help="Name of the new solver block")
    parser.add_argument("--output", type=Path, default=None, help="Output .xlsx file")
    args = parser.parse_args()

    excel_path = args.excel
    output_path = args.output or excel_path.with_name(excel_path.stem + "_with_" + args.name + ".xlsx")

    opt_results = read_results(args.opt_dir)
    dec_results = read_results(args.dec_dir)

    print(f"OPT results found: {len(opt_results)} from {args.opt_dir}")
    print(f"DEC results found: {len(dec_results)} from {args.dec_dir}")

    wb = load_workbook(excel_path)

    # Add to both full result sheets and selection sheets when present.
    summary = []
    if opt_results:
        for sheet_name in ["OPT", "Selection OPT"]:
            if sheet_name in wb.sheetnames:
                matched, col = add_or_update_block(wb[sheet_name], opt_results, "opt", args.name)
                summary.append((sheet_name, matched, get_column_letter(col)))

    if dec_results:
        for sheet_name in ["resultsDEC(2)", "Selection DEC"]:
            if sheet_name in wb.sheetnames:
                matched, col = add_or_update_block(wb[sheet_name], dec_results, "dec", args.name)
                summary.append((sheet_name, matched, get_column_letter(col)))

    wb.save(output_path)

    print(f"Saved: {output_path}")
    for sheet_name, matched, col_letter in summary:
        print(f"  {sheet_name}: matched {matched} rows, block starts at column {col_letter}")


if __name__ == "__main__":
    main()
